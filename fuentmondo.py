from datetime import datetime
import re
import subprocess
import hashlib
import requests
import json
import copy
import openpyxl
import os
import io
import base64
import msal
import webbrowser

import pyperclip
import time
import sys
from collections import defaultdict
from dotenv import load_dotenv
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart


load_dotenv()

CLIENT_ID = os.getenv("CLIENT_ID")
GRAPH_API_ENDPOINT = 'https://graph.microsoft.com/v1.0'
AUTHORITY = 'https://login.microsoftonline.com/common/'
SCOPES = ['Files.ReadWrite.All']
ONEDRIVE_SHARE_LINK = "https://1drv.ms/x/s!AidvQapyuNp6jBKR5uMUCaBYdLl0?e=3kXyKW"
SANCIONES_FILE = "sanciones.json"
VIOLACIONES_FILE = "violaciones.json"
TEAMS_FILE = "teams.json"

# Globales para control de caché
FORCE_REFRESH = False
FORCED_ROUND_IDS = set()


# Muestra un menú en la terminal para que el usuario elija el modo de ejecución.
def choose_save_option():
    print("\n--- MODO DE EJECUCIÓN ---")
    print("1. Actualizar Excel en OneDrive")
    print("2. Actualizar Excel Localmente")
    print("3. Generar Solo Informe (HTML)")
    print("4. [Opcional] Forzar Refresco de Datos (Ignorar Caché)")

    force_refresh = False
    while True:
        choice = input("\nElige una opción (1-4): ").strip()
        if choice == '4':
            force_refresh = True
            print(">>> Caché desactivada para esta ejecución. <<<")
            print("\nAhora elige el modo de guardado:")
            print("1. Actualizar Excel en OneDrive")
            print("2. Actualizar Excel Localmente")
            print("3. Generar Solo Informe (HTML)")
            continue

        if choice == '1':
            return 'onedrive', force_refresh
        elif choice == '2':
            return 'local', force_refresh
        elif choice == '3':
            return 'multas_only', force_refresh
        else:
            print("Opción no válida. Por favor, introduce 1, 2, 3 o 4.")

# Muestra el código de autenticación en la terminal.
def show_auth_code_window(message, verification_uri):
    try:
        user_code = message.split("enter the code ")[1].split(" to authenticate")[0]
    except IndexError:
        user_code = "No se pudo extraer el código"

    print("\n" + "="*60)
    print("AUTENTICACIÓN REQUERIDA")
    print("="*60)
    print(f"1. Copia este código: {user_code}")
    print(f"2. Abre esta URL en tu navegador: {verification_uri}")
    print("="*60)

    pyperclip.copy(user_code)
    print("(El código ha sido copiado al portapapeles automáticamente)")

    webbrowser.open(verification_uri)
    input("\nPresiona Enter después de haberte autenticado en el navegador...")

# Se autentica de forma interactiva y obtiene un token de acceso para Microsoft Graph.
def get_access_token():
    app = msal.PublicClientApplication(CLIENT_ID, authority=AUTHORITY)
    result = None
    accounts = app.get_accounts()
    if accounts:
        result = app.acquire_token_silent(SCOPES, account=accounts[0])
    if not result:
        flow = app.initiate_device_flow(scopes=SCOPES)
        if "error" in flow:
            print(f"\nERROR AL INICIAR LA AUTENTICACIÓN:\nError: {flow.get('error')}\nDescripción: {flow.get('error_description')}")
            return None
        show_auth_code_window(flow["message"], flow["verification_uri"])
        result = app.acquire_token_by_device_flow(flow)
    if "access_token" in result:
        return result['access_token']
    else:
        print("Error al obtener el token de acceso:", result.get("error_description"))
        return None

# Codifica un enlace de compartición de OneDrive a un formato compatible con la API de Graph.
def encode_sharing_link(sharing_link):
    base64_value = base64.b64encode(sharing_link.encode('utf-8')).decode('utf-8')
    return 'u!' + base64_value.rstrip('=').replace('/', '_').replace('+', '-')

# Obtiene el ID del Drive y el ID del archivo a partir de un enlace de compartición.
def get_drive_item_from_share_link(access_token, share_url):
    encoded_url = encode_sharing_link(share_url)
    api_url = f"{GRAPH_API_ENDPOINT}/shares/{encoded_url}/driveItem"
    headers = {'Authorization': f'Bearer {access_token}'}
    response = requests.get(api_url, headers=headers)
    response.raise_for_status()
    data = response.json()
    return data['parentReference']['driveId'], data['id']

# Descarga el contenido de un archivo Excel desde OneDrive.
def download_excel_from_onedrive(access_token, drive_id, item_id):
    api_url = f"{GRAPH_API_ENDPOINT}/drives/{drive_id}/items/{item_id}/content"
    headers = {'Authorization': f'Bearer {access_token}'}
    response = requests.get(api_url, headers=headers)
    response.raise_for_status()
    print("Excel descargado de OneDrive con éxito.")
    return response.content

# Sube (o sobrescribe) el contenido de un archivo Excel a OneDrive, con reintentos si está bloqueado.
def upload_excel_to_onedrive(access_token, drive_id, item_id, file_content):
    api_url = f"{GRAPH_API_ENDPOINT}/drives/{drive_id}/items/{item_id}/content"
    headers = {
        'Authorization': f'Bearer {access_token}',
        'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    }
    max_retries = 3
    retry_delay = 5
    for attempt in range(max_retries):
        try:
            response = requests.put(api_url, headers=headers, data=file_content)
            response.raise_for_status()
            print("Excel subido a OneDrive con éxito.")
            return
        except requests.exceptions.HTTPError as e:
            if e.response.status_code == 423 and attempt < max_retries - 1:
                print(f"El archivo está bloqueado. Reintentando en {retry_delay} segundos... (Intento {attempt + 1}/{max_retries})")
                time.sleep(retry_delay)
            else:
                raise
    print("No se pudo subir el archivo después de varios intentos.")

# Carga un archivo JSON (payload) desde una ruta específica.
def cargar_payload(ruta_archivo):
    try:
        with open(ruta_archivo, 'r', encoding='utf-8') as archivo:
            return json.load(archivo)
    except (FileNotFoundError, json.JSONDecodeError) as e:
        print(f"Error al cargar '{ruta_archivo}': {e}")
        return None

# Realiza una llamada POST a una API con un payload JSON.
def _llamar_api_raw(url, payload):
    if not payload: return None
    try:
        response = requests.post(url, json=payload)
        response.raise_for_status()
        return response.json()
    except requests.exceptions.RequestException as e:
        print(f"Error en la llamada a la API '{url}': {e}")
        return None

def llamar_api(url, payload, use_cache=True, force_refresh=None):
    if force_refresh is None:
        force_refresh = FORCE_REFRESH

    # Forzar refresco para jornadas marcadas (última o especiales) o metadatos críticos
    if not force_refresh and use_cache:
        rid = payload.get('query', {}).get('round') or payload.get('query', {}).get('roundNumber')
        if rid and rid in FORCED_ROUND_IDS:
            force_refresh = True
        elif any(endpoint in url for endpoint in ["/ranking/general", "/userteam/rounds"]):
            # Forzar también rankings generales y lista de rondas para detectar cambios
            force_refresh = True

    if not use_cache:
        return _llamar_api_raw(url, payload)

    # Crear identificador único para la petición
    payload_str = json.dumps(payload, sort_keys=True)
    query_hash = hashlib.md5(f"{url}{payload_str}".encode()).hexdigest()
    cache_dir = ".cache"
    cache_file = os.path.join(cache_dir, f"{query_hash}.json")

    # El cache expira tras 10 minutos
    CACHE_EXPIRATION = 600

    if not force_refresh and os.path.exists(cache_file):
        try:
            mtime = os.path.getmtime(cache_file)
            if time.time() - mtime < CACHE_EXPIRATION:
                with open(cache_file, 'r', encoding='utf-8') as f:
                    time.sleep(0.01) # Pequeño retardo para visibilidad
                    return json.load(f)
        except Exception:
            pass

    # Si no está en caché, ha expirado, o se fuerza el refresco, llamar a la API
    data = _llamar_api_raw(url, payload)

    if data and 'answer' in data and data['answer'] != 'api.error.general':
        try:
            os.makedirs(cache_dir, exist_ok=True)
            with open(cache_file, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False)
        except Exception as e:
            print(f"Error al guardar cache: {e}")

    return data

def print_progress(current, total, prefix='', suffix='', decimals=1, length=50, fill='█', print_end="\r"):
    if total == 0:
        print(f"{prefix} No hay elementos para procesar. {suffix}")
        return
    percent = ("{0:." + str(decimals) + "f}").format(100 * (current / float(total)))
    filled_length = int(length * current // total)
    bar = fill * filled_length + '-' * (length - filled_length)
    print(f'\r{prefix} |{bar}| {percent}% {suffix}', end=print_end)
    if current == total:
        print()

# Guarda los datos de respuesta de la API en un archivo JSON.
def guardar_respuesta(datos, nombre_archivo):
    try:
        os.makedirs(os.path.dirname(nombre_archivo), exist_ok=True)
        with open(nombre_archivo, 'w', encoding='utf-8') as f:
            json.dump(datos, f, indent=4, ensure_ascii=False)
        print(f"Respuesta de la API guardada en '{nombre_archivo}'.")
    except Exception as e:
        print(f"Error al guardar el archivo '{nombre_archivo}': {e}")

# Obtiene una lista de los capitanes de todos los equipos para una ronda específica.
def get_captains_for_round(payload_base, datos_ronda, name_map={}):
    team_captains = []
    if 'answer' not in datos_ronda or 'ranking' not in datos_ronda['answer']:
        return []
    ranking = datos_ronda['answer']['ranking']
    for team_info in ranking:
        team_id = team_info['_id']
        team_name_api = team_info['name']

        lineup_players = get_lineup_for_round(payload_base, datos_ronda['query']['roundNumber'], team_id)

        if lineup_players:
            capitan = next((p['name'] for p in lineup_players if p.get('cpt')), "N/A")
            canonical_name = name_map.get(team_name_api, team_name_api)
            team_captains.append({"team_name": canonical_name, "capitan": capitan})

    return team_captains

# Obtiene la alineación de un equipo para una ronda específica.
def get_lineup_for_round(payload_base, round_id, team_id):
    if str(round_id).startswith("LOCAL_"):
        return []
    API_URL_LINEUP = "https://api.futmondo.com/1/userteam/roundlineup"
    payload_lineup = {
        "header": copy.deepcopy(payload_base["header"]),
        "query": {
            "championshipId": payload_base["query"]["championshipId"],
            "round": round_id,
            "userteamId": team_id
        }
    }
    datos_lineup = llamar_api(API_URL_LINEUP, payload_lineup)
    if datos_lineup and 'answer' in datos_lineup and 'players' in datos_lineup['answer']:
        return datos_lineup['answer']['players']
    return []

# Procesa la respuesta de la API de rondas, manejando números de ronda enteros y flotantes.
def procesar_rondas_api(rounds_list):
    if not rounds_list:
        return {}

    rounds_map = {}
    all_numbers = [r['number'] for r in rounds_list if r.get('number') is not None]
    existing_integer_rounds = {int(n) for n in all_numbers if n % 1 == 0}

    for r in rounds_list:
        round_num = r.get('number')
        round_id = r.get('id')

        if round_num is None or not round_id:
            continue

        if round_num % 1 == 0:
            rounds_map[int(round_num)] = round_id
        else:
            truncated_num = int(round_num)
            if truncated_num in existing_integer_rounds:
                rounds_map[round_num] = round_id
            else:
                rounds_map[truncated_num] = round_id

    return rounds_map

# Calcula las multas de una jornada con un desglose detallado.
def calcular_multas_jornada(teams_in_round, matches, team_map_name, dict_alineaciones, dict_capitanes, lista_peores_equipos, peores_jugadores_final, peores_capitanes_final):
    multas_finales = {}
    for team_name in teams_in_round:
        multas_finales[team_name] = {
            "multa_total": 0.0,
            "desglose": {
                "jugadores_repetidos": {"cantidad": 0, "multa": 0.0},
                "capitan_repetido_con_rival": {"aplicado": False, "multa": 0.0},
                "tenias_capitan_rival": {"aplicado": False, "multa": 0.0},
                "peor_equipo_jornada": {"posicion": 0, "multa": 0.0},
                "alinear_peor_jugador": {"aplicado": False, "multa": 0.0},
                "elegir_peor_capitan": {"aplicado": False, "multa": 0.0},
                "alineacion_indebida": {"cantidad": 0, "multa": 0.0, "jugadores": []}
            }
        }
    for match in matches:
        team_indices = match['p']
        team_a_name = team_map_name.get(team_indices[0])
        team_b_name = team_map_name.get(team_indices[1])
        if not team_a_name or not team_b_name:
            continue

        lineup_a = dict_alineaciones.get(team_a_name, [])
        lineup_b = dict_alineaciones.get(team_b_name, [])
        nombres_a = {p['name'] for p in lineup_a}
        nombres_b = {p['name'] for p in lineup_b}
        capitan_a = dict_capitanes.get(team_a_name, "N/A")
        capitan_b = dict_capitanes.get(team_b_name, "N/A")

        repetidos_iniciales = nombres_a.intersection(nombres_b)
        repetidos_para_multa = repetidos_iniciales.copy()

        if capitan_a != "N/A":
            repetidos_para_multa.discard(capitan_a)
        if capitan_b != "N/A":
            repetidos_para_multa.discard(capitan_b)

        if repetidos_para_multa:
            cantidad_repetidos_multables = len(repetidos_para_multa)
            multa_repetidos = cantidad_repetidos_multables * 0.5
            multas_finales[team_a_name]["desglose"]["jugadores_repetidos"] = {"cantidad": cantidad_repetidos_multables, "multa": multa_repetidos}
            multas_finales[team_b_name]["desglose"]["jugadores_repetidos"] = {"cantidad": cantidad_repetidos_multables, "multa": multa_repetidos}

        if capitan_a == capitan_b and capitan_a != "N/A":
            multas_finales[team_a_name]["desglose"]["capitan_repetido_con_rival"] = {"aplicado": True, "multa": 1.0}
            multas_finales[team_b_name]["desglose"]["capitan_repetido_con_rival"] = {"aplicado": True, "multa": 1.0}
        if capitan_a in nombres_b and capitan_a != capitan_b:
            multas_finales[team_b_name]["desglose"]["tenias_capitan_rival"] = {"aplicado": True, "multa": 1.0}
        if capitan_b in nombres_a and capitan_a != capitan_b:
            multas_finales[team_a_name]["desglose"]["tenias_capitan_rival"] = {"aplicado": True, "multa": 1.0}

    multas_peores = {1: 2.0, 2: 1.5, 3: 1.0}
    for item in lista_peores_equipos:
        pos = item['posicion']
        team_name = item['equipo']
        if team_name in multas_finales and pos in multas_peores:
            multas_finales[team_name]["desglose"]["peor_equipo_jornada"] = {"posicion": pos, "multa": multas_peores[pos]}

    nombres_peores_jugadores = {p['nombre'] for p in peores_jugadores_final}
    for team_name, alineacion in dict_alineaciones.items():
        nombres_jugadores = {p['name'] for p in alineacion}
        if not nombres_peores_jugadores.isdisjoint(nombres_jugadores):
            multas_finales[team_name]["desglose"]["alinear_peor_jugador"] = {"aplicado": True, "multa": 1.0}

    nombres_peores_capitanes = {p['nombre'] for p in peores_capitanes_final}
    for team_name, capitan in dict_capitanes.items():
        if capitan in nombres_peores_capitanes:
            multas_finales[team_name]["desglose"]["elegir_peor_capitan"] = {"aplicado": True, "multa": 1.0}

    for team_name, data in multas_finales.items():
        total = sum(d.get('multa', 0.0) for d in data['desglose'].values())
        multas_finales[team_name]['multa_total'] = round(total, 2)

    return multas_finales

# Carga el archivo de sanciones o devuelve una estructura vacía si no existe.
def cargar_sanciones(ruta_archivo):
    data = {"primera": {}, "segunda": {}}
    if os.path.exists(ruta_archivo):
        try:
            with open(ruta_archivo, 'r', encoding='utf-8') as f:
                loaded = json.load(f)
                if isinstance(loaded, dict):
                    data.update(loaded)
        except (FileNotFoundError, json.JSONDecodeError):
            pass
    return data

# Guarda el estado actual de las sanciones en un archivo JSON.
def guardar_sanciones(datos, ruta_archivo):
    try:
        with open(ruta_archivo, 'w', encoding='utf-8') as f:
            json.dump(datos, f, indent=4, ensure_ascii=False)
        print(f"Archivo de sanciones guardado en '{ruta_archivo}'.")
    except Exception as e:
        print(f"Error al guardar el archivo de sanciones: {e}")

# Carga el archivo de violaciones o devuelve una estructura vacía si no existe.
def cargar_violaciones(ruta_archivo):
    data = {"primera": {}, "segunda": {}}
    if os.path.exists(ruta_archivo):
        try:
            with open(ruta_archivo, 'r', encoding='utf-8') as f:
                loaded = json.load(f)
                if isinstance(loaded, dict):
                    data.update(loaded)
        except (FileNotFoundError, json.JSONDecodeError):
            pass
    return data

# Guarda el estado actual de las violaciones en un archivo JSON.
def guardar_violaciones(datos, ruta_archivo):
    try:
        with open(ruta_archivo, 'w', encoding='utf-8') as f:
            json.dump(datos, f, indent=4, ensure_ascii=False)
        print(f"Archivo de violaciones guardado en '{ruta_archivo}'.")
    except Exception as e:
        print(f"Error al guardar el archivo de violaciones: {e}")

# Carga el archivo de equipos o devuelve una estructura vacía si no existe.
def cargar_equipos(ruta_archivo):
    if os.path.exists(ruta_archivo):
        try:
            with open(ruta_archivo, 'r', encoding='utf-8') as f:
                data = json.load(f)
                if isinstance(data, dict):
                    return data
        except Exception as e:
            print(f"Error al cargar {ruta_archivo}: {e}")
    return {"primera": {}, "segunda": {}}

# Guarda la configuración de equipos en teams.json.
def guardar_equipos(datos, ruta_archivo):
    try:
        with open(ruta_archivo, 'w', encoding='utf-8') as f:
            json.dump(datos, f, ensure_ascii=False, indent=4)
        print(f"Archivo de equipos guardado en '{ruta_archivo}'.")
    except Exception as e:
        print(f"Error al guardar el archivo de equipos: {e}")

# Envía un correo con *todas* las sanciones activas y añade CC en Lunes/Viernes.
def enviar_correo_sanciones(sanciones_por_division, violaciones_detectadas=None, force_send=False):
    EMAIL_HOST = os.getenv("EMAIL_HOST")
    EMAIL_PORT = os.getenv("EMAIL_PORT")
    EMAIL_USER = os.getenv("EMAIL_USER")
    EMAIL_PASSWORD = os.getenv("EMAIL_PASSWORD")
    EMAIL_RECIPIENT = os.getenv("EMAIL_RECIPIENT")
    # Nueva variable de entorno para destinatarios en copia
    EMAIL_RECIPIENTS_CC = os.getenv("EMAIL_RECIPIENTS_CC") or os.getenv("EMAIL_RECIPIENT_CC")

    if not all([EMAIL_HOST, EMAIL_PORT, EMAIL_USER, EMAIL_PASSWORD, EMAIL_RECIPIENT]):
        print("Faltan variables de entorno para el envío de correo. No se enviará la notificación.")
        return

    cuerpo_mensaje = ""
    hay_sanciones_activas = False

    for division, equipos in sanciones_por_division.items():
        titulo_division = "1ª DIVISIÓN" if division == "primera" else "2ª DIVISIÓN"
        buffer_division = ""
        division_con_sanciones = False

        for equipo, jugadores in sorted(equipos.items()):
            buffer_equipo = ""
            equipo_con_sanciones_en_buffer = False

            for jugador, sanciones in sorted(jugadores.items()):
                sancion_a_mostrar = next((s for s in sanciones if s.get('status') == 'active'), None)
                if not sancion_a_mostrar:
                    sancion_a_mostrar = next((s for s in sanciones if s.get('status') == 'captain_banned'), None)

                if not sancion_a_mostrar:
                    continue

                hay_sanciones_activas = True
                division_con_sanciones = True
                equipo_con_sanciones_en_buffer = True
                estado_str = ""

                if sancion_a_mostrar['status'] == 'active':
                    restantes = sancion_a_mostrar.get('games_to_serve', 3) - sancion_a_mostrar.get('games_served', 0)
                    if restantes > 0:
                        estado_str = f"Sancionado, le queda(n) {restantes} partido(s) por cumplir."
                    else:
                        estado_str = "Puede volver al once, pero no como capitan."
                elif sancion_a_mostrar['status'] == 'captain_banned':
                    jornada_fin_restriccion = sancion_a_mostrar.get('jornada_completed', 0) + 3
                    estado_str = f"Puede volver al once, pero no como capitan. Restricción hasta la Jornada {jornada_fin_restriccion}."

                buffer_equipo += f"{jugador}: {estado_str}\n"

            if equipo_con_sanciones_en_buffer:
                buffer_division += f"- {equipo}\n{buffer_equipo}"

        if division_con_sanciones:
            cuerpo_mensaje += f"\n*Sanciones Activas - {titulo_division}*\n{buffer_division}"

    if violaciones_detectadas:
        cuerpo_mensaje += "\n\n⚠️  *ALINEACIONES INDEBIDAS DETECTADAS (MULTA 5€)* ⚠️\n"
        hay_sanciones_activas = True # Forzar envío si hay multas nuevas
        for division, violaciones in violaciones_detectadas.items():
            if not violaciones: continue
            titulo_division = "1ª DIVISIÓN" if division == "primera" else "2ª DIVISIÓN"
            cuerpo_mensaje += f"\n{titulo_division}:\n"
            for team_name, lista_multas in violaciones.items():
                for m in lista_multas:
                    cuerpo_mensaje += f"- {team_name}: Alineó a {m['jugador']} en la Jornada {m['jornada']} (Sancionado).\n"

    if not hay_sanciones_activas:
        if force_send:
            cuerpo_mensaje = "No se han detectado nuevas sanciones ni alineaciones indebidas en esta jornada.\n"
        else:
            print("No se detectaron sanciones activas. No se enviará correo.")
            return

    cuerpo_mensaje += "\n\nInfo completa en blackmanx.github.io/fuentmondo"
    msg = MIMEMultipart()
    msg['From'] = EMAIL_USER
    msg['To'] = EMAIL_RECIPIENT
    msg['Subject'] = f"Informe de Sanciones SuperLiga - {datetime.now().strftime('%d/%m/%Y')}"
    msg.attach(MIMEText(cuerpo_mensaje, 'plain'))

    # --- Lógica de CC para Lunes y Viernes ---
    dia_semana = datetime.now().weekday() # Lunes es 0, Viernes es 4
    lista_destinatarios = [EMAIL_RECIPIENT]
    mensaje_cc = ""

    if EMAIL_RECIPIENTS_CC and dia_semana in [1, 4]:
        msg['Cc'] = EMAIL_RECIPIENTS_CC
        # Asumimos que EMAIL_RECIPIENTS_CC es una cadena de correos separados por coma
        lista_destinatarios.extend([email.strip() for email in EMAIL_RECIPIENTS_CC.split(',')])
        mensaje_cc = f" (y a destinatarios en copia: {EMAIL_RECIPIENTS_CC})"
        print(f"Hoy es martes o viernes. Añadiendo destinatarios en CC: {EMAIL_RECIPIENTS_CC}")
    # --- Fin de la lógica de CC ---

    try:
        server = smtplib.SMTP(EMAIL_HOST, int(EMAIL_PORT))
        server.starttls()
        server.login(EMAIL_USER, EMAIL_PASSWORD)
        text = msg.as_string()
        # Usamos la lista de destinatarios (To + Cc) para el método sendmail
        server.sendmail(EMAIL_USER, lista_destinatarios, text)
        server.quit()
        print(f"Correo de informe de sanciones enviado a '{EMAIL_RECIPIENT}'{mensaje_cc}.")
    except Exception as e:
        print(f"Error al enviar el correo de notificación: {e}")

# Genera el HTML para la tabla de multas de una jornada.
def _generar_tabla_multas_jornada_html(multas_data):
    sorted_teams = sorted(multas_data.items(), key=lambda item: item[1]['multa_total'], reverse=True)
    table_rows = ""
    for i, (team_name, data) in enumerate(sorted_teams):
        multa_total = data.get('multa_total', 0.0)
        if multa_total == 0: continue
        desglose = data.get('desglose', {})
        desglose_html = "<ul class='list-disc list-inside space-y-1'>"

        jr = desglose.get("jugadores_repetidos", {})
        if jr.get("multa", 0) > 0:
            desglose_html += f"<li>Jugadores repetidos ({jr.get('cantidad', 0)}): {jr.get('multa', 0):.2f}€</li>"
        cr = desglose.get("capitan_repetido_con_rival", {})
        if cr.get("multa", 0) > 0:
            desglose_html += f"<li>Capitán repetido con rival: {cr.get('multa', 0):.2f}€</li>"
        tcr = desglose.get("tenias_capitan_rival", {})
        if tcr.get("multa", 0) > 0:
            desglose_html += f"<li>Alinear al capitán del rival: {tcr.get('multa', 0):.2f}€</li>"
        pe = desglose.get("peor_equipo_jornada", {})
        if pe.get("multa", 0) > 0:
            pos_map = {1: "Peor", 2: "2º Peor", 3: "3er Peor"}
            pos_str = pos_map.get(pe.get("posicion"), f"{pe.get('posicion')}º Peor")
            desglose_html += f"<li>{pos_str} equipo de la jornada: {pe.get('multa', 0):.2f}€</li>"
        apj = desglose.get("alinear_peor_jugador", {})
        if apj.get("multa", 0) > 0:
            desglose_html += f"<li>Alinear al peor jugador: {apj.get('multa', 0):.2f}€</li>"
        epc = desglose.get("elegir_peor_capitan", {})
        if epc.get("multa", 0) > 0:
            desglose_html += f"<li>Elegir al peor capitán: {epc.get('multa', 0):.2f}€</li>"

        ali = desglose.get("alineacion_indebida", {})
        if ali.get("multa", 0) > 0:
            jugadores_str = ", ".join(ali.get("jugadores", []))
            desglose_html += f"<li class='text-red-600 font-bold'>Alineación indebida ({jugadores_str}): {ali.get('multa', 0):.2f}€</li>"

        desglose_html += "</ul>"

        row_bg = 'bg-slate-50' if i % 2 != 0 else 'bg-white'
        table_rows += f"""
        <tr class="{row_bg}">
            <td class="p-3 border border-slate-300">{team_name}</td>
            <td class="p-3 border border-slate-300 text-center font-bold text-red-600">{multa_total:.2f}€</td>
            <td class="p-3 border border-slate-300">{desglose_html}</td>
        </tr>"""
    if not table_rows:
        table_rows = '<tr><td colspan="3" class="text-center p-4 border border-slate-300">No se registraron multas en esta jornada.</td></tr>'

    return f"""
    <div class="overflow-x-auto">
        <table class="w-full text-left border-collapse">
            <thead class="bg-slate-200">
                <tr>
                    <th class="p-3 font-bold uppercase text-slate-600 border border-slate-300">Equipo</th>
                    <th class="p-3 font-bold uppercase text-slate-600 border border-slate-300 text-center">Multa Total</th>
                    <th class="p-3 font-bold uppercase text-slate-600 border border-slate-300">Desglose</th>
                </tr>
            </thead>
            <tbody>{table_rows}</tbody>
        </table>
    </div>"""

# Genera el HTML para la tabla de multas totales acumuladas.
def _generar_tabla_multas_totales_html(multas_acumuladas):
    sorted_teams = sorted(multas_acumuladas.items(), key=lambda item: item[1], reverse=True)
    table_rows = ""
    for i, (team_name, total_multa) in enumerate(sorted_teams):
        row_bg = 'bg-slate-50' if i % 2 != 0 else 'bg-white'
        table_rows += f"""
        <tr class="{row_bg}">
            <td class="p-3 border border-slate-300">{team_name}</td>
            <td class="p-3 border border-slate-300 text-center font-bold text-red-600">{total_multa:.2f}€</td>
        </tr>"""
    return f"""
    <div class="overflow-x-auto">
        <table class="w-full text-left border-collapse">
            <thead class="bg-slate-200">
                <tr>
                    <th class="p-3 font-bold uppercase text-slate-600 border border-slate-300">Equipo</th>
                    <th class="p-3 font-bold uppercase text-slate-600 border border-slate-300 text-center">Total Acumulado</th>
                </tr>
            </thead>
            <tbody>{table_rows}</tbody>
        </table>
    </div>"""

# Genera el HTML para la tabla de clasificación.
def _generar_tabla_clasificacion_html(ranking_ordenado):
    table_rows = ""
    for i, equipo in enumerate(ranking_ordenado):
        row_bg = 'bg-slate-50' if i % 2 != 0 else 'bg-white'

        comentario_html = ""
        if equipo.get('comentario'):
             comentario_html = f"<div class='mt-1 text-xs text-red-600 font-semibold italic'>{equipo['comentario']}</div>"

        table_rows += f"""
        <tr class="{row_bg}">
            <td class="p-3 border border-slate-300 text-center">{i + 1}</td>
            <td class="p-3 border border-slate-300">
                <div class="font-medium">{equipo['name']}</div>
                {comentario_html}
            </td>
            <td class="p-3 border border-slate-300 text-center font-bold">{equipo['points']}</td>
            <td class="p-3 border border-slate-300 text-center text-slate-500">{equipo['general_points']}</td>
        </tr>"""
    return f"""
    <div class="overflow-x-auto rounded-lg shadow-sm">
        <table class="w-full text-left border-collapse min-w-[600px]">
            <thead class="bg-slate-200 text-slate-700">
                <tr>
                    <th class="p-3 font-bold uppercase text-xs tracking-wider border border-slate-300 text-center w-12">Pos.</th>
                    <th class="p-3 font-bold uppercase text-xs tracking-wider border border-slate-300">Equipo</th>
                    <th class="p-3 font-bold uppercase text-xs tracking-wider border border-slate-300 text-center w-24">Puntos (J)</th>
                    <th class="p-3 font-bold uppercase text-xs tracking-wider border border-slate-300 text-center w-24">Puntos (G)</th>
                </tr>
            </thead>
            <tbody>{table_rows}</tbody>
        </table>
    </div>"""

# Genera el HTML para la tabla de historial de capitanes.
def _generar_tabla_capitanes_html(capitanes_por_jornada, lista_equipos, range_rounds=None):
    if not capitanes_por_jornada:
        return "<p class='text-slate-500 italic'>No hay datos de capitanes registrados.</p>"

    # Filtrar jornadas si se especifica un rango
    all_j_nums = sorted(capitanes_por_jornada.keys(), reverse=True)
    if range_rounds:
        sorted_jornadas = [j for j in all_j_nums if j in range_rounds]
    else:
        sorted_jornadas = all_j_nums

    if not sorted_jornadas:
        return "<p class='text-slate-500 italic text-sm text-center p-4 border border-dashed border-slate-300 rounded'>Sin registros en este periodo.</p>"

    # Preparar cabecera con nombres de equipos
    # Ajuste responsive: Jornada más estrecha en móvil
    table_header = '<th class="p-2 md:p-3 font-bold uppercase text-[9px] md:text-[10px] tracking-wider border border-slate-300 bg-slate-200 sticky left-0 z-30 w-16 md:w-24 text-center">Jornada</th>'
    sorted_teams = sorted(list(lista_equipos))
    for team in sorted_teams:
        table_header += f'<th class="p-1 md:p-2 font-bold uppercase text-[9px] md:text-[10px] tracking-wider border border-slate-300 min-w-[90px] md:min-w-[120px] text-center whitespace-nowrap">{team}</th>'

    table_rows = ""
    for i, j_num in enumerate(sorted_jornadas):
        row_bg = 'bg-slate-50' if i % 2 != 0 else 'bg-white'
        # Sticky left para la columna de jornada con fondo explícito para evitar transparencia
        row_content = f'<td class="p-1 md:p-2 border border-slate-300 font-bold {row_bg} sticky left-0 z-20 text-center text-[10px] md:text-sm">{j_num}</td>'
        capitanes_j = {item['team_name']: item for item in capitanes_por_jornada[j_num]}

        for team in sorted_teams:
            info = capitanes_j.get(team)
            if info:
                # Estilo si tiene roja (sanción por 3 amarillas/capitanías)
                cell_class = "bg-yellow-100 text-yellow-800 font-bold" if info.get('is_red_card') else "text-slate-600"
                row_content += f'<td class="p-1 md:p-2 border border-slate-300 text-center text-[9px] md:text-[11px] {cell_class}">{info["capitan"]}</td>'
            else:
                row_content += f'<td class="p-1 md:p-2 border border-slate-300 bg-slate-100/30 text-slate-400 text-center">-</td>'

        table_rows += f'<tr class="{row_bg} hover:bg-slate-100 transition-colors">{row_content}</tr>'

    return f"""
    <div class="overflow-x-auto rounded-lg shadow-sm border border-slate-300 max-h-[700px] overflow-y-auto">
        <table class="w-full text-left border-collapse table-auto">
            <thead class="bg-slate-100 text-slate-700 sticky top-0 z-40">
                <tr>{table_header}</tr>
            </thead>
            <tbody class="divide-y divide-slate-200">{table_rows}</tbody>
        </table>
    </div>"""

# Genera el HTML para la tabla de sanciones.
def _generar_tabla_sanciones_html(sanciones_division, violaciones_division=None):
    if not any(sanciones_division.values()) and not violaciones_division:
        return "<p>No hay sanciones activas o recientes en esta división.</p>"

    table_rows = ""
    equipos_con_sanciones = {team: players for team, players in sanciones_division.items() if any(s.get('status') != 'completed' for p in players.values() for s in p)}

    # --- Sección de Alineaciones Indebidas ---
    if violaciones_division:
        for team_name, lista_multas in violaciones_division.items():
            for m in lista_multas:
                 table_rows += f"""
                <tr class="bg-red-50">
                    <td class="p-3 border border-slate-300 font-bold text-red-700">{team_name}</td>
                    <td class="p-3 border border-slate-300 font-bold text-red-700">{m['jugador']}</td>
                    <td class="p-3 border border-slate-300 font-bold text-red-700">
                        ALINEACIÓN INDEBIDA (Jornada {m['jornada']})<br>
                        Multa: 5€
                    </td>
                </tr>"""

    if not equipos_con_sanciones and not violaciones_division:
        return "<p>No hay sanciones activas o recientes en esta división.</p>"

    sorted_teams = sorted(equipos_con_sanciones.items())

    row_index = 0
    for team_name, jugadores in sorted_teams:
        sorted_jugadores = sorted(jugadores.items())

        for player_name, sanciones in sorted_jugadores:
            sancion_a_mostrar = next((s for s in sanciones if s.get('status') == 'active'), None)
            if not sancion_a_mostrar:
                sancion_a_mostrar = next((s for s in sanciones if s.get('status') == 'captain_banned'), None)

            if not sancion_a_mostrar:
                continue

            row_bg = 'bg-slate-50' if row_index % 2 != 0 else 'bg-white'
            estado_html = ""

            if sancion_a_mostrar['status'] == 'active':
                restantes = sancion_a_mostrar.get('games_to_serve', 3) - sancion_a_mostrar.get('games_served', 0)
                if restantes > 0:
                    estado_html = f"<span class='font-bold text-red-600'>Sancionado</span><br>Le quedan {restantes} partido(s) por cumplir."
                else:
                    estado_html = "<span class='font-semibold text-orange-500'>Puede volver al once, pero no como capitan.</span>"

            elif sancion_a_mostrar['status'] == 'captain_banned':
                jornada_fin_restriccion = sancion_a_mostrar.get('jornada_completed', 0) + 3
                estado_html = f"<span class='font-semibold text-orange-500'>Puede volver al once, pero no como capitan.</span><br>Esta restricción dura hasta la Jornada {jornada_fin_restriccion}."

            table_rows += f"""
            <tr class="{row_bg}">
                <td class="p-3 border border-slate-300">{team_name}</td>
                <td class="p-3 border border-slate-300">{player_name}</td>
                <td class="p-3 border border-slate-300">{estado_html}</td>
            </tr>"""
            row_index += 1

    if not table_rows:
        return "<p>No hay sanciones activas o recientes en esta división.</p>"

    return f"""
    <div class="overflow-x-auto rounded-lg shadow-sm">
        <table class="w-full text-left border-collapse min-w-[600px]">
            <thead class="bg-slate-200 text-slate-700">
                <tr>
                    <th class="p-3 font-bold uppercase text-xs tracking-wider border border-slate-300">Equipo</th>
                    <th class="p-3 font-bold uppercase text-xs tracking-wider border border-slate-300">Jugador</th>
                    <th class="p-3 font-bold uppercase text-xs tracking-wider border border-slate-300">Estado de la Sanción</th>
                </tr>
            </thead>
            <tbody>{table_rows}</tbody>
        </table>
    </div>"""

# Genera el HTML para la tabla de historial de alineaciones indebidas.
def _generar_tabla_violaciones_html(violaciones_division):
    if not violaciones_division:
        return "<p class='text-slate-500 italic'>No hay alineaciones indebidas registradas.</p>"

    table_rows = ""
    # Aplanar la estructura: lista de (team, violacion)
    lista_plana = []
    for team_name, lista_multas in violaciones_division.items():
        for m in lista_multas:
            lista_plana.append({'team': team_name, **m})

    # Ordenar por jornada descendente
    lista_plana.sort(key=lambda x: x['jornada'], reverse=True)

    for item in lista_plana:
        table_rows += f"""
        <tr class="bg-red-50 hover:bg-red-100 transition-colores">
            <td class="p-3 border border-red-200 font-bold text-red-800 text-center">{item['jornada']}</td>
            <td class="p-3 border border-red-200 font-bold text-red-800">{item['team']}</td>
            <td class="p-3 border border-red-200 text-red-700">{item['jugador']}</td>
            <td class="p-3 border border-red-200 text-red-700 text-center font-mono font-bold">5.00€</td>
        </tr>"""

    return f"""
    <div class="overflow-x-auto rounded-lg shadow-sm border border-red-200">
        <table class="w-full text-left border-collapse min-w-[600px]">
            <thead class="bg-red-100 text-red-800">
                <tr>
                    <th class="p-3 font-bold uppercase text-xs tracking-wider border border-red-200 text-center w-20">Jornada</th>
                    <th class="p-3 font-bold uppercase text-xs tracking-wider border border-red-200">Equipo</th>
                    <th class="p-3 font-bold uppercase text-xs tracking-wider border border-red-200">Jugador</th>
                    <th class="p-3 font-bold uppercase text-xs tracking-wider border border-red-200 text-center w-24">Multa</th>
                </tr>
            </thead>
            <tbody>{table_rows}</tbody>
        </table>
    </div>"""

# Genera la página HTML completa con todos los datos y la navegación usando Tailwind CSS.
def generar_pagina_html_completa(datos_informe, output_path, current_matchday=None):
    contenido_html = ""
    nav_links_html = ""

    for div_key, div_data in datos_informe.items():
        div_titulo = "1ª División" if div_key == "primera" else "2ª División"

        id_clasificacion = f"{div_key}-clasificacion"
        id_sanciones = f"{div_key}-sanciones"
        id_violaciones = f"{div_key}-violaciones"
        id_capitanes = f"{div_key}-capitanes"
        id_totales = f"{div_key}-totales"

        nav_links_html += f'<a href="#" class="nav-link block px-4 py-2 text-white hover:bg-slate-700 md:inline-block rounded-md transition-colors" data-target="{id_clasificacion}">Clasificación {div_titulo}</a>'
        nav_links_html += f'<a href="#" class="nav-link block px-4 py-2 text-white hover:bg-slate-700 md:inline-block rounded-md transition-colors" data-target="{id_sanciones}">Sanciones {div_titulo}</a>'
        nav_links_html += f'<a href="#" class="nav-link block px-4 py-2 text-white hover:bg-slate-700 md:inline-block rounded-md transition-colors" data-target="{id_violaciones}">Alineaciones Indebidas {div_titulo}</a>'
        nav_links_html += f'<a href="#" class="nav-link block px-4 py-2 text-white hover:bg-slate-700 md:inline-block rounded-md transition-colors" data-target="{id_capitanes}">Historial Capitanes {div_titulo}</a>'
        nav_links_html += f'<a href="#" class="nav-link block px-4 py-2 text-white hover:bg-slate-700 md:inline-block rounded-md transition-colors" data-target="{id_totales}">Multas Totales {div_titulo}</a>'

        contenido_html += f'<div id="{id_clasificacion}" class="content-section p-4 md:p-6 bg-white rounded-lg shadow-md mb-6"> <h2 class="text-xl md:text-2xl font-bold text-center text-slate-700 border-b-2 border-sky-500 pb-3 mb-6">Clasificación - {div_titulo}</h2> {_generar_tabla_clasificacion_html(div_data["clasificacion"])} </div>'
        contenido_html += f'<div id="{id_sanciones}" class="content-section p-4 md:p-6 bg-white rounded-lg shadow-md mb-6"> <h2 class="text-xl md:text-2xl font-bold text-center text-slate-700 border-b-2 border-sky-500 pb-3 mb-6">Sanciones Activas - {div_titulo}</h2> {_generar_tabla_sanciones_html(div_data["sanciones"])} </div>'
        contenido_html += f'<div id="{id_violaciones}" class="content-section p-4 md:p-6 bg-white rounded-lg shadow-md mb-6"> <h2 class="text-xl md:text-2xl font-bold text-center text-red-700 border-b-2 border-red-500 pb-3 mb-6">Historial Alineaciones Indebidas - {div_titulo}</h2> {_generar_tabla_violaciones_html(div_data.get("violaciones_historico"))} </div>'

        # Tablas de capitanes por temporadas (mismo vista)
        rounds_new = range(20, 100)
        rounds_old = range(1, 20)

        # Identificar equipos para cada periodo
        equipos_new = set()
        equipos_old = set()
        for r_num, items in div_data["capitanes"].items():
            if r_num in rounds_new:
                for item in items: equipos_new.add(item['team_name'])
            if r_num in rounds_old:
                for item in items: equipos_old.add(item['team_name'])

        contenido_html += f'<div id="{id_capitanes}" class="content-section p-4 md:p-6 bg-white rounded-lg shadow-md mb-6 space-y-8">'
        contenido_html += f'<h2 class="text-xl md:text-2xl font-bold text-center text-slate-700 border-b-2 border-sky-500 pb-3 mb-2">Historial Capitanes - {div_titulo}</h2>'

        # Tabla Temporada Actual
        if equipos_new:
            contenido_html += f'<div><h3 class="text-lg font-semibold text-slate-600 mb-3 ml-1">Temporada Actual (20+)</h3>{_generar_tabla_capitanes_html(div_data["capitanes"], sorted(list(equipos_new)), rounds_new)}</div>'

        # Tabla Temporada Invierno
        if equipos_old:
            contenido_html += f'<div class="pt-4"><h3 class="text-lg font-semibold text-slate-600 mb-3 ml-1">Temporada Invierno (1-19)</h3>{_generar_tabla_capitanes_html(div_data["capitanes"], sorted(list(equipos_old)), rounds_old)}</div>'

        contenido_html += '</div>'

        contenido_html += f'<div id="{id_totales}" class="content-section p-4 md:p-6 bg-white rounded-lg shadow-md mb-6"> <h2 class="text-xl md:text-2xl font-bold text-center text-slate-700 border-b-2 border-sky-500 pb-3 mb-6">Multas Totales - {div_titulo}</h2> {_generar_tabla_multas_totales_html(div_data["totales"])} </div>'

        # Dropdown solo para las multas por jornada
        nav_links_html += '<div class="relative dropdown-container">'
        nav_links_html += f'<button class="dropdown-btn block w-full text-left px-4 py-2 text-white hover:bg-slate-700 md:inline-block md:w-auto rounded-md transition-colors">Multas Jornada ({div_titulo}) &#9662;</button>'
        nav_links_html += '<div class="dropdown-content hidden md:absolute bg-white text-black rounded-md shadow-lg mt-2 py-1 z-20 w-full md:w-48 max-h-64 overflow-y-auto">'

        sorted_jornadas = sorted(div_data['jornadas'], key=lambda x: x['numero'])
        for jornada_data in sorted_jornadas:
            jornada_num = jornada_data['numero']
            id_jornada = f"{div_key}-jornada-{jornada_num}"
            contenido_html += f'<div id="{id_jornada}" class="content-section hidden p-4 md:p-6 bg-white rounded-lg shadow-md mb-6"> <h2 class="text-xl md:text-2xl font-bold text-center text-slate-700 border-b-2 border-sky-500 pb-3 mb-6">Multas Jornada {jornada_num} - {div_titulo}</h2> {_generar_tabla_multas_jornada_html(jornada_data["multas"])} </div>'
            nav_links_html += f'<a href="#" class="block px-4 py-2 hover:bg-slate-100 text-sm" data-target="{id_jornada}">Jornada {jornada_num}</a>'

        nav_links_html += '</div></div>'

    html_completo = f"""
    <!DOCTYPE html>
    <html lang="es" class="scroll-smooth">
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>Informe - SuperLiga Fuentmondo</title>
        <script src="https://cdn.jsdelivr.net/npm/@tailwindcss/browser@4"></script>
    </head>
    <body class="bg-slate-100 text-slate-800 font-sans">

        <header class="bg-slate-800 text-white flex justify-between items-center p-4 shadow-lg fixed top-0 left-0 right-0 z-50">
            <div class="flex flex-col">
                <h1 class="text-xl font-bold">Informe SuperLiga</h1>
                {f'<span class="text-sm text-slate-300">Jornada Actual: {current_matchday}</span>' if current_matchday else ''}
            </div>
            <button id="hamburger-btn" class="md:hidden text-2xl">☰</button>
            <nav id="navbar" class="fixed top-0 left-0 h-full w-64 bg-slate-800 transform -translate-x-full transition-transform duration-300 ease-in-out md:relative md:translate-x-0 md:flex md:w-auto md:h-auto md:bg-transparent">
                <div class="p-4 md:flex md:items-center md:gap-2">
                    {nav_links_html}
                </div>
            </nav>
        </header>

        <div id="overlay" class="fixed inset-0 bg-black bg-opacity-50 z-30 hidden md:hidden"></div>

        <main class="pt-20">
            <div class="max-w-7xl mx-auto px-4 sm:px-6 lg:px-8 py-8 space-y-8">
                {contenido_html}
            </div>
        </main>

        <script>
            document.addEventListener('DOMContentLoaded', () => {{
                const hamburgerBtn = document.getElementById('hamburger-btn');
                const navbar = document.getElementById('navbar');
                const overlay = document.getElementById('overlay');
                const navLinks = document.querySelectorAll('.nav-link, .dropdown-content a');
                const dropdownBtns = document.querySelectorAll('.dropdown-btn');

                function toggleMenu() {{
                    const isOffScreen = navbar.classList.contains('-translate-x-full');
                    navbar.classList.toggle('-translate-x-full', !isOffScreen);
                    navbar.classList.toggle('translate-x-0', isOffScreen);
                    overlay.classList.toggle('hidden');
                }}

                hamburgerBtn.addEventListener('click', toggleMenu);
                overlay.addEventListener('click', toggleMenu);

                function showContent(targetId) {{
                    document.querySelectorAll('.content-section').forEach(section => section.classList.add('hidden'));
                    const targetElement = document.getElementById(targetId);
                    if (targetElement) {{
                        targetElement.classList.remove('hidden');
                    }}

                    document.querySelectorAll('.nav-link').forEach(link => link.classList.remove('bg-sky-600'));
                    const activeLink = document.querySelector(`.nav-link[data-target='${{targetId}}']`);
                    if (activeLink) {{
                        activeLink.classList.add('bg-sky-600');
                    }}
                }}

                navLinks.forEach(link => {{
                    link.addEventListener('click', e => {{
                        e.preventDefault();
                        const targetId = e.currentTarget.dataset.target;
                        showContent(targetId);
                        if (window.innerWidth < 768) {{
                            toggleMenu();
                        }}
                        document.querySelectorAll('.dropdown-content').forEach(d => d.classList.add('hidden'));
                    }});
                }});

                dropdownBtns.forEach(btn => {{
                    btn.addEventListener('click', e => {{
                        e.stopPropagation();
                        const dropdownContent = e.currentTarget.nextElementSibling;
                        document.querySelectorAll('.dropdown-content').forEach(d => {{
                            if (d !== dropdownContent) d.classList.add('hidden');
                        }});
                        dropdownContent.classList.toggle('hidden');
                    }});
                }});

                window.addEventListener('click', () => {{
                    document.querySelectorAll('.dropdown-content').forEach(d => d.classList.add('hidden'));
                }});

                const firstSectionId = document.querySelector('.content-section')?.id;
                if (firstSectionId) {{
                    showContent(firstSectionId);
                }}
            }});
        </script>
    </body>
    </html>"""

    try:
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(html_completo)
        print(f"Informe HTML (Tailwind CSS) completo guardado en '{output_path}'.")
    except Exception as e:
        print(f"Error al guardar el archivo HTML final: {e}")

# Procesa y ordena los datos de clasificación de la API.
def _procesar_y_ordenar_clasificacion(datos_general, datos_teams, name_map={}):
    puntos_generales_dict = {name_map.get(e['teamname'], e['teamname']): e['points'] for e in datos_teams['answer']['teams']}
    ranking_general_list = datos_general['answer']['ranking']
    equipos_para_ordenar = []
    for equipo in ranking_general_list:
        nombre_equipo_api = equipo['name']
        nombre_equipo_canonico = name_map.get(nombre_equipo_api, nombre_equipo_api)

        puntos_general = puntos_generales_dict.get(nombre_equipo_canonico, 0)
        puntos_jornada = equipo['points']
        comentario = ""

        # Ajustes manuales de puntos
        if nombre_equipo_canonico == "EL CHOLISMO FC":
            puntos_jornada -= 3
            comentario = "Sanción: -3 puntos"
        elif nombre_equipo_canonico == "LA MARRANERA":
            puntos_jornada += 3
            comentario = "Ajuste por sanción a rival: +3 puntos"

        equipos_para_ordenar.append({
            'name': nombre_equipo_canonico,
            'points': puntos_jornada,
            'general_points': puntos_general,
            'comentario': comentario
        })
    return sorted(equipos_para_ordenar, key=lambda x: (x['points'], x['general_points']), reverse=True)

# Actualiza una hoja de Excel con los datos de clasificación.
def actualizar_hoja_excel(workbook, ranking_ordenado, sheet_name, fila_inicio, columna_inicio):
    try:
        sheet = workbook[sheet_name]
        for row in sheet.iter_rows(min_row=fila_inicio, max_row=sheet.max_row, min_col=columna_inicio, max_col=columna_inicio + 2):
            for cell in row:
                cell.value = None
        for i, equipo in enumerate(ranking_ordenado):
            fila_actual = fila_inicio + i
            sheet.cell(row=fila_actual, column=columna_inicio).value = equipo['name']
            sheet.cell(row=fila_actual, column=columna_inicio + 1).value = equipo['points']
            sheet.cell(row=fila_actual, column=columna_inicio + 2).value = equipo['general_points']
        print(f"Hoja '{sheet_name}' actualizada en memoria.")
    except Exception as e:
        print(f"Error procesando la hoja '{sheet_name}' en memoria: {e}")

# Actualiza las cabeceras con los nombres de los equipos en la hoja 'Capitanes'.
def actualizar_cabeceras_capitanes(workbook, teams_dict_1a, teams_dict_2a):
    try:
        sheet = workbook["Capitanes"]
        print("Actualizando cabeceras de 1ª División en la hoja 'Capitanes'...")
        sorted_teams_1a = sorted(teams_dict_1a.items(), key=lambda item: int(item[0]))
        current_col = 3
        for _, team_name in sorted_teams_1a:
            sheet.cell(row=3, column=current_col).value = team_name
            current_col += 2
        print("Actualizando cabeceras de 2ª División en la hoja 'Capitanes'...")
        sorted_teams_2a = sorted(teams_dict_2a.items(), key=lambda item: int(item[0]))
        current_col = 44
        for _, team_name in sorted_teams_2a:
            sheet.cell(row=3, column=current_col).value = team_name
            current_col += 2
        print("Cabeceras de la hoja 'Capitanes' actualizadas.")
    except Exception as e:
        print(f"Error al actualizar las cabeceras de la hoja 'Capitanes': {e}")

# Actualiza la fila de una jornada con los capitanes de cada equipo en el Excel.
def actualizar_hoja_capitanes(workbook, round_number, team_captains_list):
    try:
        sheet = workbook["Capitanes"]
        team_to_captain_col = {}
        for col_idx in range(3, 150):
            team_name_cell = sheet.cell(row=3, column=col_idx)
            if team_name_cell.value:
                team_name = str(team_name_cell.value).strip()
                team_to_captain_col[team_name] = col_idx
        row_found = False
        target_row_label = f"Jornada {round_number}"
        for row_idx, row in enumerate(sheet.iter_rows(min_row=5, max_row=sheet.max_row, min_col=2, max_col=2), 5):
            cell_value = row[0].value
            if isinstance(cell_value, str) and cell_value.strip().lower() == target_row_label.lower():
                for captain_info in team_captains_list:
                    team_name = captain_info['team_name'].strip()
                    captain = captain_info['capitan']
                    if team_name in team_to_captain_col:
                        col_to_update = team_to_captain_col[team_name]
                        sheet.cell(row=row_idx, column=col_to_update).value = captain
                    else:
                        print(f"     -> Aviso: El equipo '{team_name}' no se encontró en la cabecera de la hoja 'Capitanes'.")
                print(f"Capitanes de la '{target_row_label}' actualizados en memoria.")
                row_found = True
                break
        if not row_found:
            print(f"Advertencia: No se encontró la fila para '{target_row_label}' en la hoja 'Capitanes'.")
    except Exception as e:
        print(f"Error actualizando la hoja 'Capitanes' para la jornada {round_number}: {e}")

# Itera sobre todas las jornadas para actualizar el histórico de capitanes en el Excel.
def actualizar_capitanes_historico(workbook, rounds_map, payload_base, division_name, name_map={}):
    print(f"\n--- INICIANDO ACTUALIZACIÓN HISTÓRICA DE CAPITANES PARA {division_name.upper()} (EXCEL) ---")
    sorted_round_numbers = sorted(rounds_map.keys())
    total_rounds = len(sorted_round_numbers)
    for i, round_number in enumerate(sorted_round_numbers):
        round_id = rounds_map[round_number]
        print_progress(i, total_rounds, prefix=f'Jornada {round_number}:', suffix='Excel...', length=40)
        payload_round = copy.deepcopy(payload_base)
        payload_round['query'].update({'roundNumber': round_id, 'championshipId': payload_base['query']['championshipId']})
        datos_ronda = llamar_api("https://api.futmondo.com/1/ranking/round", payload_round)
        if not datos_ronda or 'answer' not in datos_ronda or datos_ronda['answer'] == 'api.error.general':
            print(f"\n     -> Error: No se pudieron obtener datos para la Jornada {round_number}. Saltando.")
            continue
        datos_ronda['query']['roundNumber'] = round_id
        team_captains = get_captains_for_round(payload_base, datos_ronda, name_map)
        if not team_captains:
            print(f"\n     -> Advertencia: No se encontraron capitanes para la Jornada {round_number}.")
            continue
        actualizar_hoja_capitanes(workbook, round_number, team_captains)
    print_progress(total_rounds, total_rounds, prefix='Finalizado:', suffix='Excel actualizado', length=40)

# Procesa todos los datos de una ronda y calcula las multas correspondientes.
def procesar_ronda_completa(datos_ronda, output_file, payload_base, name_map={}):
    if not datos_ronda or 'answer' not in datos_ronda or 'matches' not in datos_ronda['answer']:
        print("Error: Respuesta de API de ronda inválida.")
        return None
    teams_in_round_list = datos_ronda['answer'].get('ranking', [])
    matches = datos_ronda['answer']['matches']
    round_id_actual = datos_ronda['query']['roundId']
    team_map_id = {i + 1: team['_id'] for i, team in enumerate(teams_in_round_list)}
    team_map_name = {i + 1: name_map.get(team['name'], team['name']) for i, team in enumerate(teams_in_round_list)}
    resultados_finales, puntos_equipos_por_ronda, jugadores_ronda = [], [], []
    dict_alineaciones, dict_capitanes = {}, {}
    for match in matches:
        ids = [team_map_id.get(p) for p in match['p']]
        nombres = [team_map_name.get(p) for p in match['p']]
        puntos = match.get('data', {}).get('partial', match.get('m', [0, 0]))
        for i in range(2):
            if nombres[i]:
                puntos_equipos_por_ronda.append({"equipo": nombres[i], "puntos": puntos[i]})
        lineups, capitanes = [], []
        for i in range(2):
            if not ids[i]: continue
            payload_lineup = {
                "header": copy.deepcopy(payload_base["header"]),
                "query": {"championshipId": payload_base["query"]["championshipId"], "round": round_id_actual, "userteamId": ids[i]}
            }
            datos_lineup = llamar_api("https://api.futmondo.com/1/userteam/roundlineup", payload_lineup)
            lineup_players = datos_lineup.get('answer', {}).get('players', [])
            lineups.append(lineup_players)
            capitan = next((p['name'] for p in lineup_players if p.get('cpt')), "N/A")
            capitanes.append(capitan)
            if nombres[i]:
                dict_alineaciones[nombres[i]] = lineup_players
                dict_capitanes[nombres[i]] = capitan
            for player in lineup_players:
                jugadores_ronda.append({
                    "nombre": player['name'], "puntos": player['points'],
                    "equipo": nombres[i], "es_capitan": player.get('cpt', False)
                })
        jugadores_repetidos = [p['name'] for p in lineups[0] if p['name'] in {p2['name'] for p2 in lineups[1]}]
        if nombres[0] and nombres[1]:
            resultados_finales.append({
                "Combate": f"{nombres[0]} vs {nombres[1]}",
                f"{nombres[0]}": {"Puntuacion": puntos[0], "Capitan": capitanes[0]},
                f"{nombres[1]}": {"Puntuacion": puntos[1], "Capitan": capitanes[1]},
                "Jugadores repetidos": jugadores_repetidos
            })
    peores_equipos = sorted(puntos_equipos_por_ronda, key=lambda x: x['puntos'])[:3]
    lista_peores_equipos = [{"posicion": i + 1, **equipo} for i, equipo in enumerate(peores_equipos)]
    def encontrar_peores(jugadores, key_filter=None):
        min_puntos = float('inf')
        peores_map = defaultdict(list)
        iterable = filter(key_filter, jugadores) if key_filter else jugadores
        for jugador in iterable:
            puntos_jugador = jugador.get('puntos', 0)
            if puntos_jugador < min_puntos:
                min_puntos = puntos_jugador
                peores_map.clear()
            if puntos_jugador == min_puntos:
                if jugador['equipo'] and jugador['equipo'] not in peores_map[jugador['nombre']]:
                    peores_map[jugador['nombre']].append(jugador['equipo'])
        return [{"nombre": n, "puntos": min_puntos, "equipos": e} for n, e in peores_map.items()]
    peores_capitanes_final = encontrar_peores(jugadores_ronda, lambda j: j.get('es_capitan'))
    peores_jugadores_final = encontrar_peores(jugadores_ronda)
    resumen_final = {
        "Resultados por combate": resultados_finales,
        "Peor Capitan": peores_capitanes_final,
        "Peor Jugador": peores_jugadores_final,
        "Los 3 peores equipos de la ronda": lista_peores_equipos
    }
    guardar_respuesta(resumen_final, output_file)
    multas_jornada = calcular_multas_jornada(
        teams_in_round=list(team_map_name.values()), matches=matches, team_map_name=team_map_name,
        dict_alineaciones=dict_alineaciones, dict_capitanes=dict_capitanes,
        lista_peores_equipos=lista_peores_equipos, peores_jugadores_final=peores_jugadores_final,
        peores_capitanes_final=peores_capitanes_final
    )
    return multas_jornada

# Itera sobre todas las jornadas para procesar y devolver los resultados y multas.
def procesar_historico_jornadas(rounds_map, payload_base, name_map, division_str):
    print(f"\n--- RECOPILANDO DATOS DE MULTAS PARA {division_str.upper()} ---")
    multas_acumuladas = defaultdict(float)
    datos_jornadas = []
    sorted_rounds = sorted(rounds_map.keys())
    total_rounds = len(sorted_rounds)
    for i, round_number in enumerate(sorted_rounds):
        round_id = rounds_map[round_number]
        print_progress(i, total_rounds, prefix=f'Jornada {round_number}:', suffix='Completado', length=40)
        payload_round = copy.deepcopy(payload_base)
        payload_round['query'].update({'roundNumber': round_id})
        datos_ronda = llamar_api("https://api.futmondo.com/1/ranking/round", payload_round)
        if datos_ronda and 'answer' in datos_ronda:
            datos_ronda['query']['roundId'] = round_id
            output_file = f"resultados/jornada_{round_number}_{division_str}.json"
            multas_de_la_jornada = procesar_ronda_completa(datos_ronda, output_file, payload_base, name_map)
            if multas_de_la_jornada:
                datos_jornadas.append({'numero': round_number, 'multas': multas_de_la_jornada})
                for team, data in multas_de_la_jornada.items():
                    multas_acumuladas[team] += data.get('multa_total', 0.0)
        else:
            print(f"\nNo se pudieron obtener datos para la Jornada {round_number}. Saltando.")
    print_progress(total_rounds, total_rounds, prefix='Finalizado:', suffix='Completado', length=40)
    return datos_jornadas, dict(multas_acumuladas)

# Obtiene el capitán de un equipo desde un archivo local de resultados.
def _extraer_capitan_de_archivo_local(round_number, team_name):
    for div in ["primera", "segunda"]:
        local_path = f"resultados/jornada_{round_number}_{div}.json"
        if os.path.exists(local_path):
            try:
                with open(local_path, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    for match in data.get('Resultados por combate', []):
                        if team_name in match:
                            return match[team_name].get('Capitan', "N/A")
            except Exception:
                pass
    return "N/A"

# Recopila el historial de capitanes y alineaciones para procesar las sanciones de forma iterativa.
def procesar_sanciones_y_capitanes(rounds_map, payload_base, name_map, division_str, sanciones_existentes):
    print(f"\n--- PROCESANDO SANCIONES Y CAPITANES PARA {division_str.upper()} ---")

    sanciones_actualizadas = copy.deepcopy(sanciones_existentes)
    nuevas_sanciones = defaultdict(dict)
    sorted_rounds = sorted(rounds_map.keys())

    # Paso 1: Recopilar todos los datos históricos de capitanes.
    all_teams_data = {}
    print("Obteniendo datos históricos de todas las jornadas para análisis...")
    total_rounds = len(sorted_rounds)
    for i, round_number in enumerate(sorted_rounds):
        round_id = rounds_map[round_number]
        print_progress(i, total_rounds, prefix=f'Jornada {round_number}:', suffix='Analizando...', length=40)

        if str(round_id).startswith("LOCAL_"):
            # Para rondas locales, buscamos en los JSON de resultados
            for team_name in name_map.values():
                capitan = _extraer_capitan_de_archivo_local(round_number, team_name)
                # No tenemos lista de jugadores completa en local
                all_teams_data.setdefault(team_name, {})[round_number] = {
                    'capitan': capitan,
                    'players': []
                }
            continue

        payload_round = copy.deepcopy(payload_base)
        payload_round['query'].update({'roundNumber': round_id})
        datos_ronda = llamar_api("https://api.futmondo.com/1/ranking/round", payload_round)
        if not datos_ronda or 'answer' not in datos_ronda: continue

        for team_info in datos_ronda.get('answer', {}).get('ranking', []):
            team_id, team_name_api = team_info['_id'], team_info['name']
            team_name = name_map.get(team_name_api, team_name_api)

            lineup_players = get_lineup_for_round(payload_base, round_id, team_id)
            capitan = next((p['name'] for p in lineup_players if p.get('cpt')), "N/A")

            all_teams_data.setdefault(team_name, {})[round_number] = {
                'capitan': capitan,
                'players': [p['name'] for p in lineup_players]
            }
    print_progress(total_rounds, total_rounds, prefix='Finalizado:', suffix='Análisis completo', length=40)

    # Paso 2: Procesar la lógica de sanciones de forma cronológica
    contador_capitanes = defaultdict(lambda: defaultdict(int))
    multas_alineacion_indebida = defaultdict(list)

    for team_name in all_teams_data.keys():
        sanciones_actualizadas.setdefault(team_name, {})

        for round_number in sorted_rounds:
            round_data = all_teams_data.get(team_name, {}).get(round_number)
            if not round_data: continue

            # A. Actualizamos el estado de las sanciones existentes para este equipo
            for player, sanciones in sanciones_actualizadas[team_name].items():

                # Sanciones de partido (estado 'active').
                for sancion in filter(lambda s: s.get('status') == 'active', sanciones):
                    # Verificar alineación indebida
                    current_players = round_data.get('players', [])
                    if player in current_players and round_number > sancion['jornada_triggered']:
                        multas_alineacion_indebida[team_name].append({
                            'jornada': round_number,
                            'jugador': player,
                            'multa': 5.0
                        })

                    rounds_passed = round_number - sancion['jornada_triggered']

                    if rounds_passed > 0:
                        sancion['games_served'] = min(rounds_passed, sancion.get('games_to_serve', 3))

                    if sancion.get('games_served', 0) >= sancion.get('games_to_serve', 3):
                        sancion['status'] = 'captain_banned'
                        sancion['jornada_completed'] = sancion['jornada_triggered'] + sancion.get('games_to_serve', 3)

                # Sanciones de capitanía (estado 'captain_banned').
                for sancion in filter(lambda s: s.get('status') == 'captain_banned', sanciones):
                    if round_number >= sancion.get('jornada_completed', float('inf')) + 3:
                        sancion['status'] = 'completed'
                        sancion['jornada_fully_cleared'] = round_number

            # B. Verificamos si se genera una NUEVA sanción en esta jornada
            capitan = round_data['capitan']
            if capitan != "N/A":
                contador_capitanes[team_name][capitan] += 1

                if contador_capitanes[team_name][capitan] % 3 == 0:
                    sanciones_jugador = sanciones_actualizadas[team_name].setdefault(capitan, [])
                    if not any(s['status'] == 'active' for s in sanciones_jugador):
                        nueva_sancion = {
                            'type': '3_match_ban',
                            'jornada_triggered': round_number,
                            'status': 'active',
                            'games_to_serve': 3,
                            'games_served': 0
                        }
                        sanciones_jugador.append(nueva_sancion)
                        nuevas_sanciones[team_name][capitan] = nueva_sancion

    # Paso 3: Preparar los datos finales para la tabla HTML
    capitanes_para_informe = {}
    for round_number in sorted_rounds:
        capitanes_para_informe[round_number] = []
        for team_name, rounds_data in all_teams_data.items():
            if round_number in rounds_data:
                capitan_name = rounds_data[round_number]['capitan']
                cap_info = {'team_name': team_name, 'capitan': capitan_name}

                sancion_del_dia = next((s for s in sanciones_actualizadas.get(team_name, {}).get(capitan_name, []) if s.get('jornada_triggered') == round_number), None)
                if sancion_del_dia:
                    cap_info['is_red_card'] = True

                capitanes_para_informe[round_number].append(cap_info)

    return capitanes_para_informe, sanciones_actualizadas, nuevas_sanciones, multas_alineacion_indebida

# Sube el informe HTML a un repositorio de GitHub automáticamente.
def subir_informe_a_github(ruta_archivo_html, GITHUB_TOKEN, GITHUB_USERNAME, GITHUB_REPO):
    print("\n--- INTENTANDO SUBIR INFORME A GITHUB ---")
    try:
        remote_url = f"https://{GITHUB_TOKEN}@github.com/{GITHUB_USERNAME}/{GITHUB_REPO}.git"
        subprocess.run(["git", "add", ruta_archivo_html], check=True, capture_output=True, text=True)
        status_result = subprocess.run(["git", "status", "--porcelain"], check=True, capture_output=True, text=True)
        if ruta_archivo_html not in status_result.stdout:
            print("✅ No hay cambios detectados en el informe. No se necesita subir nada.")
            return
        mensaje_commit = f"Informe actualizado automáticamente - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        subprocess.run(["git", "commit", "-m", mensaje_commit], check=True, capture_output=True, text=True)
        subprocess.run(["git", "push", "--force", remote_url], check=True, capture_output=True, text=True)
        print(f"✅ Informe '{ruta_archivo_html}' subido a GitHub con éxito.")
    except FileNotFoundError:
        print("❌ Error: Git no está instalado o no se encuentra en el PATH del sistema.")
    except subprocess.CalledProcessError as e:
        print(f"❌ Error durante la ejecución de un comando de Git: {e.stderr}")
    except Exception as e:
        print(f"❌ Ocurrió un error inesperado al intentar subir a GitHub: {e}")

# Función principal que orquesta la ejecución del script.
def main():
    GITHUB_TOKEN = os.getenv("GITHUB_TOKEN")
    GITHUB_USERNAME = os.getenv("GITHUB_USERNAME")
    GITHUB_REPO = os.getenv("GITHUB_REPO")

    LOCAL_EXCEL_FILENAME = "SuperLiga Fuentmondo 25-26.xlsx"

    # Cargar equipos desde el archivo JSON
    config_equipos = cargar_equipos(TEAMS_FILE)

    force_email = False

    # Comprobar flag --email para envío directo
    if len(sys.argv) > 1 and sys.argv[1] == '--email':
        print("Modo 'Solo Email' detectado.")
        sanciones_cargadas = cargar_sanciones(SANCIONES_FILE)
        violaciones_cargadas = cargar_violaciones(VIOLACIONES_FILE)

        if not any(sanciones_cargadas.values()) and not any(violaciones_cargadas.values()):
             print("No hay datos de sanciones o violaciones guardados para enviar.")
             return

        enviar_correo_sanciones(sanciones_cargadas, violaciones_cargadas)
        return

    # Comprueba si se pasó el argumento --auto para ejecución automática
    if len(sys.argv) > 1 and sys.argv[1] == '--auto':
        modo = 'local_auto'
        force_email = True
        print("Modo automático detectado. Actualizando localmente y forzando envío de email.")
    else:
        modo, force_refresh_choice = choose_save_option()
        global FORCE_REFRESH
        FORCE_REFRESH = force_refresh_choice

    if not modo:
        print("No se seleccionó ninguna opción. Finalizando el script.")
        return



    sanciones_iniciales = cargar_sanciones(SANCIONES_FILE)

    print("\n--- OBTENIENDO DATOS DE FUTMONDO ---")
    payload_1a = cargar_payload("payload_primera.json")
    if not payload_1a: return
    payload_2a = cargar_payload("payload.json")
    if not payload_2a: return

    datos_general_1a = llamar_api("https://api.futmondo.com/1/ranking/general", copy.deepcopy(payload_1a))
    datos_general_2a = llamar_api("https://api.futmondo.com/1/ranking/general", copy.deepcopy(payload_2a))
    payload_teams_1a = copy.deepcopy(payload_1a); payload_teams_1a['query'] = {"championshipId": payload_1a["query"]["championshipId"]}
    datos_teams_1a = llamar_api("https://api.futmondo.com/2/championship/teams", payload_teams_1a)
    payload_teams_2a = copy.deepcopy(payload_2a); payload_teams_2a['query'] = {"championshipId": payload_2a["query"]["championshipId"]}
    datos_teams_2a = llamar_api("https://api.futmondo.com/2/championship/teams", payload_teams_2a)

    rounds_data_1a = llamar_api("https://api.futmondo.com/1/userteam/rounds", copy.deepcopy(payload_1a))
    rounds_data_2a = llamar_api("https://api.futmondo.com/1/userteam/rounds", copy.deepcopy(payload_2a))

    def descubrir_rondas_locales(division_str, existing_map):
        """Escanea la carpeta resultados para encontrar jornadas que no están en el map de la API."""
        try:
            import re
            files = os.listdir("resultados")
            for f in files:
                # Buscar patrón: jornada_X_division.json
                match = re.search(rf"jornada_(\d+)_{division_str}\.json", f)
                if match:
                    num = int(match.group(1))
                    if num not in existing_map:
                        # Usamos el nombre del archivo como un "pseudo-id" ya que tenemos el archivo local
                        existing_map[num] = f"LOCAL_{num}"
            return existing_map
        except Exception as e:
            print(f"Error al buscar rondas locales para {division_str}: {e}")
            return existing_map

    rounds_map_1a = procesar_rondas_api(rounds_data_1a.get('answer', []))
    rounds_map_1a = descubrir_rondas_locales("primera", rounds_map_1a)

    rounds_map_2a = procesar_rondas_api(rounds_data_2a.get('answer', []))
    rounds_map_2a = descubrir_rondas_locales("segunda", rounds_map_2a)

    # Marcar jornadas para refresco forzoso (última y especiales con decimales)
    global FORCED_ROUND_IDS
    for r_map in [rounds_map_1a, rounds_map_2a]:
        if r_map:
            # Forzar solo la última de la API (no las locales)
            api_rounds = [k for k, v in r_map.items() if not str(v).startswith("LOCAL_")]
            if api_rounds:
                FORCED_ROUND_IDS.add(r_map[max(api_rounds)])
            # Forzar jornadas especiales (ej: 1.5)
            for r_num, r_id in r_map.items():
                if r_num % 1 != 0:
                    FORCED_ROUND_IDS.add(r_id)

    if not rounds_map_1a and not rounds_map_2a:
        print("Aviso: No se encontraron rondas en ninguna división. Es normal al inicio de temporada.")
        # Aun así, intentaremos sincronizar equipos.

    # Obtener datos de equipos para sincronización (siempre disponible)
    # ... (esto ya se hizo arriba al llamar a la API)

    datos_ronda_1a = None
    if rounds_map_1a:
        latest_round_id_1a = rounds_map_1a[max(rounds_map_1a.keys())]
        payload_round_1a = copy.deepcopy(payload_1a)
        payload_round_1a['query'].update({'roundNumber': latest_round_id_1a})
        datos_ronda_1a = llamar_api("https://api.futmondo.com/1/ranking/round", payload_round_1a)

    datos_ronda_2a = None
    if rounds_map_2a:
        latest_round_id_2a = rounds_map_2a[max(rounds_map_2a.keys())]
        payload_round_2a = copy.deepcopy(payload_2a)
        payload_round_2a['query'].update({'roundNumber': latest_round_id_2a})
        datos_ronda_2a = llamar_api("https://api.futmondo.com/1/ranking/round", payload_round_2a)

    # Sincronizar equipos de forma robusta
    def sincronizar_division(div_key, datos_teams):
        existing_div = config_equipos.get(div_key, {})
        if not datos_teams or 'answer' not in datos_teams:
            return False

        teams_list = datos_teams['answer'].get('teams', [])
        if not teams_list:
            return False

        updated = False
        # Buscamos el índice más alto actual
        max_idx = 0
        for team_data in existing_div.values():
            max_idx = max(max_idx, team_data.get('index', 0))

        for t in teams_list:
            team_id = t['id'] # El ID de futmondo es inmutable
            api_name = t['teamname']

            if team_id not in existing_div:
                max_idx += 1
                existing_div[team_id] = {
                    "index": max_idx,
                    "api_name": api_name,
                    "display_name": api_name
                }
                print(f"  [+] Nuevo equipo detectado en {div_key}: {api_name} (Index: {max_idx})")
                updated = True
            else:
                # Si el nombre ha cambiado en la API, lo actualizamos pero mantenemos el display_name
                if existing_div[team_id].get('api_name') != api_name:
                    existing_div[team_id]['api_name'] = api_name
                    updated = True

        config_equipos[div_key] = existing_div
        return updated

    updated = sincronizar_division("primera", datos_teams_1a)
    updated |= sincronizar_division("segunda", datos_teams_2a)

    if updated:
        guardar_equipos(config_equipos, TEAMS_FILE)

    # Crear los diccionarios necesarios para el resto del script
    def preparar_mappings(div_key):
        div_data = config_equipos.get(div_key, {})
        # map: API_NAME (actual) -> DISPLAY_NAME (personalizado)
        name_map = {data['api_name']: data['display_name'] for data in div_data.values()}
        # TEAMS_DICT: INDEX -> DISPLAY_NAME (para Excel)
        teams_dict = {str(data['index']): data['display_name'] for data in div_data.values()}
        return name_map, teams_dict

    map_1a, TEAMS_1A = preparar_mappings("primera")
    map_2a, TEAMS_2A = preparar_mappings("segunda")

    clasificacion_1a = _procesar_y_ordenar_clasificacion(datos_general_1a, datos_teams_1a, map_1a)
    clasificacion_2a = _procesar_y_ordenar_clasificacion(datos_general_2a, datos_teams_2a, map_2a)

    if modo in ['local', 'onedrive', 'local_auto']:
        print("\n--- PROCESANDO ARCHIVO EXCEL ---")
        if all([datos_general_1a, datos_teams_1a, datos_general_2a, datos_teams_2a]):
            workbook = None
            try:
                if modo in ['local', 'local_auto']:
                    if os.path.exists(LOCAL_EXCEL_FILENAME):
                        workbook = openpyxl.load_workbook(LOCAL_EXCEL_FILENAME)
                    else:
                        print(f"⚠️ Aviso: No se encontró el archivo Excel '{LOCAL_EXCEL_FILENAME}'. Saltando actualización local.")
                elif modo == 'onedrive':
                    access_token = get_access_token()
                    if not access_token: raise Exception("No se pudo obtener el token de acceso.")
                    drive_id, item_id = get_drive_item_from_share_link(access_token, ONEDRIVE_SHARE_LINK)
                    excel_content = download_excel_from_onedrive(access_token, drive_id, item_id)
                    workbook = openpyxl.load_workbook(io.BytesIO(excel_content))

                if workbook:
                    actualizar_cabeceras_capitanes(workbook, TEAMS_1A, TEAMS_2A)
                    # actualizar_hoja_excel(workbook, clasificacion_1a, "Clasificación 1a DIV", 5, 2)
                    # actualizar_hoja_excel(workbook, clasificacion_2a, "Clasificación 2a DIV", 2, 3)
                    actualizar_capitanes_historico(workbook, rounds_map_1a, payload_1a, "1a División", map_1a)
                    actualizar_capitanes_historico(workbook, rounds_map_2a, payload_2a, "2a División", map_2a)

                    if modo in ['local', 'local_auto']:
                        workbook.save(LOCAL_EXCEL_FILENAME)
                        print(f"\nArchivo '{LOCAL_EXCEL_FILENAME}' guardado localmente.")
                    elif modo == 'onedrive':
                        buffer = io.BytesIO()
                        workbook.save(buffer)
                        upload_excel_to_onedrive(access_token, drive_id, item_id, buffer.getvalue())
            except Exception as e:
                print(f"Error durante el procesamiento del Excel: {e}")
        else:
            print("Faltan datos clave de la API para el Excel. Saltando actualización del Excel.")
    else:
        print("\n--- MODO 'SOLO INFORME' SELECCIONADO: SALTANDO PROCESO DE EXCEL ---")

    datos_jornadas_1a, totales_1a = procesar_historico_jornadas(rounds_map_1a, payload_1a, map_1a, "primera")
    datos_jornadas_2a, totales_2a = procesar_historico_jornadas(rounds_map_2a, payload_2a, map_2a, "segunda")

    capitanes_1a, sanciones_1a, nuevas_sanciones_1a, violaciones_1a = procesar_sanciones_y_capitanes(rounds_map_1a, payload_1a, map_1a, "primera", sanciones_iniciales["primera"])
    capitanes_2a, sanciones_2a, nuevas_sanciones_2a, violaciones_2a = procesar_sanciones_y_capitanes(rounds_map_2a, payload_2a, map_2a, "segunda", sanciones_iniciales["segunda"])

    # Integrar multas por alineación indebida en datos_jornadas y totales
    def integrar_violaciones(datos_jornadas, totales, violaciones):
        for team_name, lista_multas in violaciones.items():
            for multa in lista_multas:
                jornada_num = multa['jornada']
                monto = multa['multa']
                jugador = multa['jugador']

                # Actualizar totales
                totales[team_name] = totales.get(team_name, 0.0) + monto

                # Actualizar desglose jornada
                for jornada_data in datos_jornadas:
                    if jornada_data['numero'] == jornada_num:
                        if team_name in jornada_data['multas']:
                            team_data = jornada_data['multas'][team_name]
                            team_data['multa_total'] += monto
                            desglose = team_data['desglose']
                            if 'alineacion_indebida' not in desglose:
                                desglose['alineacion_indebida'] = {"cantidad": 0, "multa": 0.0, "jugadores": []}

                            desglose['alineacion_indebida']['cantidad'] += 1
                            desglose['alineacion_indebida']['multa'] += monto
                            desglose['alineacion_indebida']['jugadores'].append(jugador)
                        break

    integrar_violaciones(datos_jornadas_1a, totales_1a, violaciones_1a)
    integrar_violaciones(datos_jornadas_2a, totales_2a, violaciones_2a)

    sanciones_finales = {"primera": sanciones_1a, "segunda": sanciones_2a}
    violaciones_totales = {"primera": violaciones_1a, "segunda": violaciones_2a}
    guardar_sanciones(sanciones_finales, SANCIONES_FILE)
    guardar_violaciones(violaciones_totales, VIOLACIONES_FILE)

    if force_email:
        print("\nModo automático: Enviando informe de sanciones (siempre habilitado)...")
        enviar_correo_sanciones(sanciones_finales, violaciones_totales, force_send=True)
    else:
        nuevas_sanciones_totales = {"primera": nuevas_sanciones_1a, "segunda": nuevas_sanciones_2a}
        hay_nuevas_sanciones = any(nuevas_sanciones_totales["primera"].values()) or any(nuevas_sanciones_totales["segunda"].values())
        hay_violaciones = any(violaciones_totales["primera"].values()) or any(violaciones_totales["segunda"].values())

        if hay_nuevas_sanciones or hay_violaciones:
            while True:
                prompt_text = "\nSe han detectado nuevas sanciones o alineaciones indebidas. ¿Quieres enviar el correo con el informe completo? (s/n): "
                respuesta = input(prompt_text).lower().strip()
                if respuesta in ['s', 'si']:
                    enviar_correo_sanciones(sanciones_finales, violaciones_totales)
                    break
                elif respuesta in ['n', 'no']:
                    print("Envío de correo cancelado por el usuario.")
                    break
                else:
                    print("Respuesta no válida. Por favor, introduce 's' para sí o 'n' para no.")
        else:
            print("\nNo se detectaron nuevas sanciones, no es necesario enviar correo.")

    datos_informe_completo = {
        "primera": {
            "jornadas": datos_jornadas_1a,
            "totales": totales_1a,
            "clasificacion": clasificacion_1a,
            "capitanes": capitanes_1a,
            "sanciones": sanciones_1a,
            "violaciones": violaciones_1a,
            "violaciones_historico": violaciones_totales["primera"]
        },
        "segunda": {
            "jornadas": datos_jornadas_2a,
            "totales": totales_2a,
            "clasificacion": clasificacion_2a,
            "capitanes": capitanes_2a,
            "sanciones": sanciones_2a,
            "violaciones": violaciones_2a,
            "violaciones_historico": violaciones_totales["segunda"]
        }
    }

    current_matchday = max(max(rounds_map_1a.keys(), default=0), max(rounds_map_2a.keys(), default=0))
    generar_pagina_html_completa(datos_informe_completo, "index.html", current_matchday)

    if all([GITHUB_TOKEN, GITHUB_USERNAME, GITHUB_REPO]):
        subir_informe_a_github("index.html", GITHUB_TOKEN, GITHUB_USERNAME, GITHUB_REPO)
    else:
        print("\n--- AVISO: Faltan variables de entorno de GitHub (.env) para la subida automática. ---")

    print("\n--- Proceso completado. ---")

if __name__ == '__main__':
    main()
